
import pdfplumber
import pandas as pd
import tempfile
import os
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import tkinter as tk
from tkinter import filedialog

if os.environ.get('DISPLAY', '') == '':
    print('No display found. Using :0.0')
    os.environ['DISPLAY'] = ':0.0'


def choose_price_and_read_df():
    global price_path
    price_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    # Меняем текст кнопки
    button1.config(text="Прайс-лист выбран")

def choose_file_and_read_df():
    global price_path
    pdf_path = filedialog.askopenfilename(filetypes=[("PDF", "*.pdf")])

# pdf_path = 'input.pdf'
# price_path = 'ekf_pricelist_2025-02-27.xlsx'

    combined_df = pd.DataFrame()

    with pdfplumber.open(pdf_path) as pdf:    
        page = pdf.pages[0]
        table = page.extract_tables()
        columns_a=table[0][0]
        # Удаляем знаки переноса из каждого элемента списка
        columns_a = [text.replace('\n', ' ') for text in columns_a]

        # Проходим по каждой странице PDF
        for i, page in enumerate(pdf.pages):
            # Извлекаем таблицы с текущей страницы
            table = page.extract_tables()
            if i == 0:
                df = pd.DataFrame(table[0][1:], columns=columns_a)
                # df.to_excel(f'exchange/df_{i}.xlsx')
            else:
                table_i=table[0]
                columns = ['None']+columns_a+['None']
                # columns = columns_a
                df = pd.DataFrame(table_i, columns=columns)
                # Удаляем первую и последнюю колонку
                df = df.drop(columns=[df.columns[0], df.columns[-1]])
                # Удаляем первые две строки
                df = df.drop([0, 1])
            combined_df = pd.concat([combined_df, df], ignore_index=True)
        combined_df['№'] = pd.to_numeric(combined_df['№'], errors='coerce').astype('int64')  # Преобразуем в целые числа
        combined_df['Артикул'] = combined_df['Артикул'].str.replace('\n', '') # Удаляем переносы из артикулов
        combined_df['Номенклатура'] = combined_df['Номенклатура'].str.replace('\n', ' ') # Удаляем переносы из Номенклатура
        combined_df['Скидка по спеццене'] = pd.to_numeric(combined_df['Скидка по спеццене'], errors='coerce').astype('int64')  # Преобразуем в целые числа
        combined_df['Количество в заказе'] = combined_df['Количество в заказе'].str.replace(' ', '') # Удаляем пробелы из числовых данных
        combined_df['Количество в заказе'] = pd.to_numeric(combined_df['Количество в заказе'], errors='coerce').astype('int64')  # Преобразуем в целые числа

        price_df = pd.read_excel(price_path)
        price_df = price_df.drop(range(10))  # Удаляем строки с индексами от 0 до 9
        # Заменяем заголовок на первую строку
        price_df.columns = price_df.iloc[0]  # Делаем первую строку заголовком
        price_df = price_df[1:].reset_index(drop=True)  # Удаляем первую строку и сбрасываем индексы
        # Выбираем только столбцы 'Артикул' и 'Базовая цена' из price_df
        df_art_base_price = price_df[['Артикул', 'Базовая цена,                   с НДС']]
        df_art_base_price.rename(columns={'Базовая цена,                   с НДС': 'Базовая цена, с НДС'}, inplace=True)

        combined_df = combined_df.merge(df_art_base_price, on='Артикул', how='left')
        combined_df['Сумма БЦ, с НДС'] = combined_df['Количество в заказе']*combined_df['Базовая цена, с НДС']
        combined_df['Цена со скидкой'] = combined_df['Базовая цена, с НДС']*(1-combined_df['Скидка по спеццене']/100)
        combined_df['Сумма со скидкой'] = combined_df['Количество в заказе']*combined_df['Цена со скидкой']
        sum_base_price = round(combined_df['Сумма БЦ, с НДС'].sum(), 2)
        sum_sale_price = round(combined_df['Сумма со скидкой'].sum(), 2)

        df_sums_line = [['', '', '', '', '', 'ИТОГО:', sum_base_price, '', sum_sale_price]]
        df_sums = pd.DataFrame(df_sums_line, columns = combined_df.columns)
        combined_df = pd.concat([combined_df, df_sums], ignore_index=True)

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        temp_filename = 'commercial_offer_EKF.xlsx' #tmp_file.name  # Получаем имя временного файла
        # Используем pd.ExcelWriter для настройки форматов и ширины столбцов
        with pd.ExcelWriter(temp_filename, engine='xlsxwriter') as writer:
        # Записываем DataFrame в Excel
            combined_df.to_excel(writer, sheet_name='Sheet1', index=False)

            # Получаем объект workbook и worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Закрепляем шапку
            worksheet.freeze_panes(1, 0)

            # Задаем числовые форматы для столбцов
            format1 = workbook.add_format({'num_format': '0'})  # Формат для №, "Скидка поспеццене" и "Количество в заказе"
            format2 = workbook.add_format({'num_format': '#,##0.00'})  # Формат для Базовая цена, с НДС
            format3 = workbook.add_format({'num_format': '#,##0.00', 'bold': True})  # Формат для Сумма БЦ, с НДС и Сумма со скидкой
            format4 = workbook.add_format({'align': 'right', 'bold': True}) # Ориент по правой стороне, жирный шрифт для последней строки
            # format3 = workbook.add_format({'num_format': '0.000%'})  # Формат для Column3

            # Применяем форматы к столбцам
            worksheet.set_column('A:A', 5, format1)  # Ширина 15 и формат для Column1
            worksheet.set_column(1, 1, 22)  # Ширина и формат для Артикул
            worksheet.set_column(2, 2, 60)  # Ширина и формат для Номенклатура
            worksheet.set_column('D:D', 18, format1)  # Ширина и формат для Скидка по спеццене
            worksheet.set_column('E:E', 18, format1)  # Ширина и формат для Количество в заказе
            worksheet.set_column('F:F', 18, format2)  # Ширина и формат для Базовая цена, с НДС
            worksheet.set_column('G:G', 18, format3)  # Ширина и формат для Сумма БЦ, с НДС
            worksheet.set_column('H:H', 18, format2)  # Ширина и формат для Цена со скидкой
            worksheet.set_column('I:I', 18, format3)  # Ширина и формат для Сумма со скидкой

            # Создаём формат для ячейки с "Итого" в последней строке
            worksheet.write(combined_df.shape[0], 5, combined_df.iloc[-1, 5], format4)

        # Закрашиваем пустые ячейки (ненайденные цены по артикулу) в жёлтый цвет
        # Открываем файл с помощью openpyxl
        wb = load_workbook(temp_filename)
        ws = wb.active
        # Применяем стили к пустым ячейкам, исключая последнюю строку
        for row in range(1, ws.max_row):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        # Сохраняем изменения
        wb.save(temp_filename)

    # Открываем файл с помощью программы по умолчанию
    if os.name == 'nt':  # Для Windows
        os.startfile(temp_filename)
    elif os.name == 'posix':  # Для macOS и Linux
        subprocess.run(['open', temp_filename])  # macOS
        # subprocess.run(['xdg-open', temp_filename])  # Linux

    # Меняем текст кнопки
    button2.config(text="Информационное письмо обработано / Выбрать заново")    


# Создание главного окна
root = tk.Tk()
root.geometry("440x180")
root.title("Конвертер ИП EKF PDF_To_XLSX")

label1 = tk.Label(                      # Создание виджета метки
    text="Выберите файлы Прайс-листа и\nИнформационного письма\nдля обработки", # Задание отображаемого текста
    font=("Arial", 12, "bold"),        # Шрифт Arial, размер 16
    fg="#07e374",                       # Установка цвета текста
    bg="#2f457c",                       # Устанавка фона
    width=60,                           # Установка ширина виджета (в текстовых юнитах)
    height=4                           # Установка высоты виджета (в текстовых юнитах) 
)
label1.pack()  

# Кнопка выбора прайс-листа
button1 = tk.Button(text="Выбрать прайс-лист", font=("Arial", 12), command=choose_price_and_read_df)
button1.pack(pady=7)

# Кнопка выбора КП
button2 = tk.Button(text="Выбрать информационное письмо", font=("Arial", 12), command=choose_file_and_read_df)
button2.pack(pady=7)

root.mainloop()